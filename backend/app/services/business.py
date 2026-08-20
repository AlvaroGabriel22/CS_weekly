import csv
import json
import logging
from datetime import UTC, date, datetime, time
from pathlib import Path

from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.activity_directives import (
    activity_requests_image_analysis,
    parse_activity_directives,
)
from app.core.exceptions import NotFoundError, QWIException, ValidationError
from app.models import (
    Activity,
    ActivityMetadata,
    ActivityStatus,
    Attachment,
    ImageUsage,
    Language,
    QualitySector,
    User,
    WeeklyReport,
    WeeklyStatus,
    WritingProfile,
)
from app.schemas.weekly import CoverageMetrics, DashboardStats
from app.schemas.weekly_content import parse_weekly_content, weekly_content_to_legacy_dict
from app.services.llm_service import LLMService
from app.services.prompt_composer import PromptComposer
from app.services.text_sanitize import sanitize_weekly_content

logger = logging.getLogger(__name__)
settings = get_settings()


def get_week_info(dt: datetime | None = None) -> tuple[int, int]:
    """(week, year) na convenção da empresa — ver app.core.dates."""
    from app.core.dates import calculate_week_number

    dt = dt or datetime.now(UTC)
    return calculate_week_number(dt)


def _layout_uses_ai(layout: dict | None) -> bool:
    """True se algum elemento do layout tem `binding` (texto preenchido pela IA).

    Um deck sem nenhum binding é 100% manual e não precisa do LLM (QA-037).
    """
    if not isinstance(layout, dict):
        return False
    for slide in layout.get("slides", []):
        for element in slide.get("elements", []):
            if isinstance(element, dict) and element.get("binding"):
                return True
    return False


class ActivityService:
    def __init__(self, db: Session, llm: LLMService | None = None):
        self.db = db
        self.llm = llm or LLMService()
        self.prompt_composer = PromptComposer()

    def create(self, user: User, data: dict) -> Activity:
        activity_date = data.get("activity_date") or datetime.now(UTC)
        week_number, year = get_week_info(activity_date)

        activity = Activity(
            user_id=user.id,
            title=data["title"],
            description=data.get("description"),
            project=None,
            category=None,
            department=user.department,
            activity_date=activity_date,
            tags=[],
            notes=None,
            include_in_weekly=True,
            week_number=week_number,
            year=year,
        )
        self.db.add(activity)
        self.db.commit()
        self.db.refresh(activity)

        return activity

    def get_by_id(self, activity_id: str, user_id: str) -> Activity:
        activity = (
            self.db.query(Activity)
            .filter(Activity.id == activity_id, Activity.user_id == user_id)
            .first()
        )
        if not activity:
            raise NotFoundError("Atividade")
        return activity

    def list_activities(
        self,
        user_id: str,
        week_number: int | None = None,
        year: int | None = None,
        start_date: date | None = None,
        end_date: date | None = None,
        timezone: str | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> tuple[list[Activity], int]:
        query = self.db.query(Activity).filter(Activity.user_id == user_id)
        if week_number:
            query = query.filter(Activity.week_number == week_number)
        if year:
            query = query.filter(Activity.year == year)
        if start_date:
            query = query.filter(Activity.activity_date >= start_date)
        if end_date:
            query = query.filter(Activity.activity_date <= end_date)

        total = query.count()
        activities = (
            query.order_by(Activity.activity_date.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        return activities, total

    def update(self, activity: Activity, data: dict) -> Activity:
        for key, value in data.items():
            if key in data and hasattr(activity, key):
                if value is not None or isinstance(value, bool):
                    setattr(activity, key, value)

        if "activity_date" in data and data["activity_date"]:
            week_number, year = get_week_info(data["activity_date"])
            activity.week_number = week_number
            activity.year = year

        activity.updated_at = datetime.now(UTC)
        self.db.commit()
        self.db.refresh(activity)
        return activity

    def delete(self, activity: Activity) -> None:
        # Remove os arquivos físicos dos anexos antes de apagar as linhas —
        # o cascade do banco não limpa o disco (QA-004).
        import shutil

        activity_dir = Path(settings.UPLOAD_DIR) / str(activity.user_id) / str(activity.id)
        self.db.delete(activity)
        self.db.commit()
        try:
            if activity_dir.exists():
                shutil.rmtree(activity_dir, ignore_errors=True)
        except Exception:
            logger.warning("Falha ao remover arquivos da atividade %s", activity.id)

    async def process_activity_metadata(self, activity: Activity) -> None:
        """Process activity through AI to extract structured metadata."""
        try:
            prompt = self.prompt_composer.compose_activity_analysis_prompt(
                activity.title, activity.description or "", activity.tags
            )
            result = await self.llm.analyze_activity(prompt)
            metadata_dict = self._parse_metadata(result)
            metadata = activity.metadata_entry or ActivityMetadata(
                activity_id=activity.id
            )
            for key, value in metadata_dict.items():
                setattr(metadata, key, value)
            metadata.processed_at = datetime.now(UTC)
            if not activity.metadata_entry:
                self.db.add(metadata)

            activity.project = metadata.project
            activity.category = metadata.category
            activity.tags = metadata.keywords or []
            activity.status = ActivityStatus.PROCESSED
            self.db.commit()
            logger.info("Activity metadata processed | activity_id=%s", activity.id)
        except Exception as e:
            logger.warning("Failed to process activity metadata | error=%s", str(e))
            metadata = activity.metadata_entry or ActivityMetadata(
                activity_id=activity.id
            )
            metadata.technical_summary = activity.description or activity.title
            if not activity.metadata_entry:
                self.db.add(metadata)
            self.db.commit()

    def _parse_metadata(self, llm_result: str) -> dict:
        try:
            start = llm_result.find("{")
            end = llm_result.rfind("}") + 1
            if start >= 0 and end > start:
                parsed = json.loads(llm_result[start:end])
                related_kpis = parsed.get("related_kpis", [])
                keywords = parsed.get("keywords", [])
                return {
                    "project": parsed.get("project"),
                    "supplier": parsed.get("supplier"),
                    "line": parsed.get("line"),
                    "process": parsed.get("process"),
                    "product": parsed.get("product"),
                    "category": parsed.get("category"),
                    "activity_type": parsed.get("activity_type"),
                    "defect_type": parsed.get("defect_type"),
                    "related_kpis": related_kpis if isinstance(related_kpis, list) else [],
                    "keywords": keywords if isinstance(keywords, list) else [],
                    "technical_summary": parsed.get("technical_summary"),
                }
        except (json.JSONDecodeError, ValueError):
            pass
        return {
            "related_kpis": [],
            "keywords": [],
            "technical_summary": llm_result[:500] if llm_result else None
        }


class FileService:
    ALLOWED_EXTENSIONS = {
        "image": {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp"},
        "spreadsheet": {".xlsx", ".xls", ".csv"},
        "document": {".pdf", ".doc", ".docx", ".pptx", ".ppt", ".txt"},
    }

    def __init__(self, db: Session, llm: LLMService | None = None):
        self.db = db
        self.llm = llm or LLMService()
        self.upload_dir = Path(settings.UPLOAD_DIR)
        self.upload_dir.mkdir(parents=True, exist_ok=True)

    def get_file_type(self, filename: str) -> str:
        ext = Path(filename).suffix.lower()
        for file_type, extensions in self.ALLOWED_EXTENSIONS.items():
            if ext in extensions:
                return file_type
        return "other"

    async def save_attachment(
        self,
        activity: Activity,
        stored_filename: str,
        original_filename: str,
        content: bytes,
        mime_type: str | None = None,
        image_usage: str | None = None,
        manual_caption: str | None = None,
        include_in_weekly: bool = True,
    ) -> Attachment:
        user_dir = self.upload_dir / activity.user_id / activity.id
        user_dir.mkdir(parents=True, exist_ok=True)

        file_path = user_dir / stored_filename
        file_path.write_bytes(content)

        resolved_image_usage = image_usage
        file_type = self.get_file_type(original_filename)
        if file_type == "image" and resolved_image_usage is None:
            resolved_image_usage = ImageUsage.INSERT_REPORT

        # Planilhas: extrai a tabela AGORA (determinístico, sem IA) para os
        # bloquinhos do editor de montagem e para a renderização no PPTX.
        kpi_data = None
        if file_type == "spreadsheet":
            from app.services.table_extract import extract_table

            table = extract_table(content, original_filename)
            if table:
                kpi_data = {"table": table}

        attachment = Attachment(
            activity_id=activity.id,
            filename=stored_filename,
            original_filename=original_filename,
            file_path=str(file_path),
            file_type=file_type,
            file_size=len(content),
            mime_type=mime_type,
            image_usage=resolved_image_usage,
            manual_caption=manual_caption,
            include_in_weekly=include_in_weekly,
            kpi_data=kpi_data,
        )
        self.db.add(attachment)
        self.db.commit()
        self.db.refresh(attachment)
        return attachment

    def analyze_spreadsheet(self, attachment: Attachment) -> dict:
        """Extract KPIs and trends from spreadsheet files."""
        try:
            sheets_data = {}
            if Path(attachment.file_path).suffix.lower() == ".csv":
                with open(
                    attachment.file_path, encoding="utf-8-sig", errors="replace"
                ) as csv_file:
                    rows = [
                        row
                        for index, row in enumerate(csv.reader(csv_file))
                        if index < 50
                    ]
                sheets_data["CSV"] = rows
            else:
                from openpyxl import load_workbook

                workbook = load_workbook(
                    attachment.file_path, read_only=True, data_only=True
                )
                for sheet_name in workbook.sheetnames[:5]:
                    sheet = workbook[sheet_name]
                    rows = []
                    for index, row in enumerate(sheet.iter_rows(values_only=True)):
                        if index >= 50:
                            break
                        rows.append(
                            [str(cell) if cell is not None else "" for cell in row]
                        )
                    sheets_data[sheet_name] = rows
                workbook.close()

            summary = json.dumps(sheets_data, ensure_ascii=False)[:3000]
            kpi_data = {
                "sheets": list(sheets_data.keys()),
                "preview": summary,
                "rows_previewed": sum(len(rows) for rows in sheets_data.values()),
            }
            # Preserva a tabela extraída no upload (usada pelo editor e pelo PPTX).
            if isinstance(attachment.kpi_data, dict) and attachment.kpi_data.get("table"):
                kpi_data["table"] = attachment.kpi_data["table"]
            attachment.kpi_data = kpi_data
            self.db.commit()
            return kpi_data
        except Exception as e:
            logger.warning("Spreadsheet analysis failed | error=%s", str(e))
            return {}

    def extract_document_text(self, attachment: Attachment) -> str:
        path = Path(attachment.file_path)
        suffix = path.suffix.lower()
        try:
            if suffix == ".txt":
                return path.read_text(encoding="utf-8", errors="replace")[:12000]
            if suffix == ".pdf":
                from pypdf import PdfReader

                reader = PdfReader(str(path))
                return "\n".join(
                    page.extract_text() or "" for page in reader.pages[:20]
                )[:12000]
            if suffix == ".docx":
                from docx import Document

                document = Document(str(path))
                return "\n".join(
                    paragraph.text for paragraph in document.paragraphs
                )[:12000]
        except Exception as error:
            logger.warning(
                "Document extraction failed | attachment_id=%s | error=%s",
                attachment.id,
                error,
            )
        return ""

    async def process_attachment(self, attachment: Attachment) -> None:
        activity = attachment.activity
        context = (
            f"Activity: {activity.title}\n"
            f"Description: {activity.description or 'Not provided'}\n"
            f"File: {attachment.original_filename}"
        )
        try:
            if attachment.file_type == "image":
                if not activity_requests_image_analysis(activity):
                    return
                raw = await self.llm.analyze_image(attachment.file_path, context)
                analysis = self._parse_json(raw)
                attachment.ai_analysis = analysis or {"raw": raw[:2000]}
                attachment.ai_caption = (
                    analysis.get("caption") if analysis else raw[:500]
                )
            elif attachment.file_type == "spreadsheet":
                local_data = self.analyze_spreadsheet(attachment)
                raw = await self.llm.analyze_spreadsheet(
                    json.dumps(local_data, ensure_ascii=False)
                )
                analysis = self._parse_json(raw)
                attachment.ai_analysis = analysis or {"raw": raw[:2000]}
                if analysis:
                    attachment.kpi_data = {
                        **(attachment.kpi_data or {}),
                        **analysis,
                    }
            elif attachment.file_type == "document":
                extracted = self.extract_document_text(attachment)
                if extracted:
                    from app.services.llm_service import QUALITY_DEPT_CONTEXT

                    response = await self.llm.generate(
                        f"{context}\n\nDocument content:\n{extracted}",
                        f"You are a quality analyst. {QUALITY_DEPT_CONTEXT} "
                        "Analyze this corporate evidence document like an analyst preparing "
                        "a management briefing: extract concrete facts, figures, dates, "
                        "conclusions and open issues. Return JSON with summary (analytical, "
                        "2-3 sentences), relevant_facts (array, each citing specifics), "
                        "possible_impacts (array), and kpis (array of indicators with values "
                        "when present). Never invent.",
                        json_mode=True,
                    )
                    parsed = self._parse_json(response.content)
                    attachment.ai_analysis = parsed or {
                        "summary": response.content[:2000]
                    }
            self.db.commit()
        except Exception as error:
            logger.warning(
                "Attachment AI processing failed | attachment_id=%s | error=%s",
                attachment.id,
                error,
            )
            if attachment.file_type == "spreadsheet" and not attachment.kpi_data:
                self.analyze_spreadsheet(attachment)
            attachment.ai_analysis = attachment.ai_analysis or {
                "status": "pending_ai",
                "message": "Stored for analysis during weekly generation.",
            }
            self.db.commit()

    def _parse_json(self, value: str) -> dict:
        try:
            start = value.find("{")
            end = value.rfind("}") + 1
            if start >= 0 and end > start:
                parsed = json.loads(value[start:end])
                return parsed if isinstance(parsed, dict) else {}
        except (json.JSONDecodeError, ValueError):
            pass
        return {}


class WeeklyService:
    def __init__(self, db: Session, llm: LLMService | None = None):
        self.db = db
        self.llm = llm or LLMService()
        self.prompt_composer = PromptComposer()

    def get_dashboard_stats(self, user: User) -> DashboardStats:
        week_number, year = get_week_info()

        activities = (
            self.db.query(Activity)
            .filter(
                Activity.user_id == user.id,
                Activity.week_number == week_number,
                Activity.year == year,
            )
            .all()
        )

        days_filled = len({a.activity_date.date() for a in activities})
        images_count = 0
        spreadsheets_count = 0
        files_count = 0

        for activity in activities:
            for att in activity.attachments:
                files_count += 1
                if att.file_type == "image":
                    images_count += 1
                elif att.file_type == "spreadsheet":
                    spreadsheets_count += 1

        weekly = (
            self.db.query(WeeklyReport)
            .filter(
                WeeklyReport.user_id == user.id,
                WeeklyReport.week_number == week_number,
                WeeklyReport.year == year,
            )
            .order_by(WeeklyReport.version.desc())
            .first()
        )

        coverage_score = 0.0
        if weekly and weekly.coverage:
            coverage_score = weekly.coverage.get("quality_score", 0.0)

        return DashboardStats(
            week_number=week_number,
            year=year,
            activities_count=len(activities),
            days_filled=days_filled,
            images_count=images_count,
            spreadsheets_count=spreadsheets_count,
            files_count=files_count,
            weekly_status=weekly.status if weekly else None,
            last_report_generated_at=weekly.generated_at if weekly else None,
            coverage_score=coverage_score,
        )

    def _generate_by_mutation(
        self,
        report: WeeklyReport,
        user: User,
        activities: list[Activity],
        pptx_template_id: str,
        *,
        week_number: int,
        year: int,
        period_label: str | None,
    ) -> WeeklyReport | None:
        """Gera o deck mutando o .pptx que o usuário enviou como modelo.

        Devolve None quando o caminho não se aplica (modelo inexistente, sem
        campos marcados, arquivo sumido) — aí a geração segue pelo fluxo antigo
        em vez de falhar. Só o que der certo aqui pula o LLM.
        """
        from app.models import PptxTemplate
        from app.services.deck_content import activities_to_content
        from app.services.deck_plan import build_plan
        from app.services.pptx_render import render_plan

        template = (
            self.db.query(PptxTemplate)
            .filter(
                PptxTemplate.id == pptx_template_id,
                PptxTemplate.user_id == user.id,
            )
            .first()
        )
        if template is None:
            logger.warning("Modelo de PPT %s não é do usuário — fluxo antigo",
                           pptx_template_id)
            return None
        layout = template.layout if isinstance(template.layout, dict) else {}
        if not layout.get("slides"):
            logger.warning("Modelo de PPT %s sem slides — fluxo antigo", template.id)
            return None

        week_label = f"W{week_number}" + (f" · {period_label}" if period_label else "")
        destino = (
            Path(get_settings().UPLOAD_DIR) / "reports" / f"{report.id}.pptx"
        )
        try:
            plan = build_plan(
                layout,
                activities_to_content(activities),
                title=report.title,
                subtitle=f"{period_label or ''} · {user.name}".strip(" ·"),
                week_label=week_label,
            )
            resultado = render_plan(template.file_path, layout, plan, destino)
        except Exception as error:
            # Modelo estranho ou arquivo corrompido não pode impedir o weekly:
            # cai no gerador antigo, que não depende do arquivo do usuário.
            logger.warning(
                "Exportação por mutação falhou (segue pelo fluxo antigo) | %s", error
            )
            return None

        report.content = {
            "raw": "",
            "structured": {},
            "ai_degraded": False,
            "layout": layout,
            "source": "pptx_mutate",
            "pptx_template_id": template.id,
            "warnings": resultado.warnings,
            "period": {
                "start": None,
                "end": None,
            },
            "activity_ids": [activity.id for activity in activities],
        }
        report.pptx_path = resultado.path
        report.prompt_used = None
        report.ai_summary = None
        report.coverage = None
        report.confidence_index = None
        report.quality_score = None
        report.status = WeeklyStatus.COMPLETED
        report.generated_at = datetime.now(UTC)
        for activity in activities:
            activity.status = ActivityStatus.USED_IN_REPORT

        # O perfil de conhecimento continua aprendendo; o de ESTILO não, porque
        # aqui não houve montagem — o layout é o do modelo, não uma escolha.
        from app.services.knowledge_profile import learn_from_activities
        learn_from_activities(self.db, user.id, activities)

        self.db.commit()
        self.db.refresh(report)
        logger.info(
            "Weekly gerado por mutação | report_id=%s | week=%d/%d | slides=%d | avisos=%d",
            report.id, week_number, year, resultado.slides, len(resultado.warnings),
        )
        return report

    async def generate_weekly(
        self,
        user: User,
        activity_ids: list[str],
        start_date: date | None = None,
        end_date: date | None = None,
        week_number: int | None = None,
        year: int | None = None,
        template_id: str | None = None,
        language: Language | None = None,
        timezone: str | None = None,
        layout: dict | None = None,
        layout_source: str = "manual",
        pptx_template_id: str | None = None,
    ) -> WeeklyReport:
        if start_date and end_date and start_date > end_date:
            start_date, end_date = end_date, start_date

        period_end = end_date or datetime.now(UTC).date()
        week_number = week_number or get_week_info(
            datetime.combine(period_end, time.min, tzinfo=UTC)
        )[0]
        year = year or get_week_info(
            datetime.combine(period_end, time.min, tzinfo=UTC)
        )[1]

        activities, _ = ActivityService(self.db).list_activities(
            user.id,
            start_date=start_date,
            end_date=end_date,
            timezone=timezone,
            page_size=500,
        )

        selected_ids = set(activity_ids)
        activities = [activity for activity in activities if activity.id in selected_ids]

        if not activities:
            raise ValidationError(
                "Nenhuma atividade válida selecionada para o período informado"
            )
        if len(activities) != len(selected_ids):
            raise ValidationError(
                "Algumas atividades selecionadas não pertencem ao período informado"
            )

        lang = language or (user.writing_profile.default_language if user.writing_profile else Language.PT)

        existing = (
            self.db.query(WeeklyReport)
            .filter(
                WeeklyReport.user_id == user.id,
                WeeklyReport.week_number == week_number,
                WeeklyReport.year == year,
            )
            .order_by(WeeklyReport.version.desc())
            .first()
        )
        version = (existing.version + 1) if existing else 1

        # Duas gerações concorrentes da mesma semana disputam a mesma `version`
        # e colidem no UNIQUE(user,year,week,version). Re-tenta com a próxima
        # versão em vez de estourar 500 (QA-042).
        from sqlalchemy.exc import IntegrityError

        report = None
        for _ in range(5):
            candidate = WeeklyReport(
                user_id=user.id,
                template_id=None,
                week_number=week_number,
                year=year,
                language=lang,
                version=version,
                status=WeeklyStatus.GENERATING,
                title=self._build_report_title(start_date, end_date, week_number, year, lang),
            )
            self.db.add(candidate)
            try:
                self.db.commit()
                report = candidate
                break
            except IntegrityError:
                self.db.rollback()
                latest = (
                    self.db.query(WeeklyReport.version)
                    .filter(
                        WeeklyReport.user_id == user.id,
                        WeeklyReport.week_number == week_number,
                        WeeklyReport.year == year,
                    )
                    .order_by(WeeklyReport.version.desc())
                    .first()
                )
                version = (latest[0] + 1) if latest else version + 1
        if report is None:
            raise QWIException(
                "Não foi possível gerar o relatório agora (concorrência). Tente novamente.",
                409,
            )

        try:
            await self._ensure_attachments_analyzed(activities)

            period_label = None
            if start_date and end_date:
                period_label = (
                    f"{start_date.strftime('%d/%m')}–{end_date.strftime('%d/%m')}"
                )

            # Modelo é um PPT enviado pelo usuário: o deck sai MUTANDO o arquivo
            # dele, não reconstruindo um novo. Caminho determinístico e sem LLM
            # — a IA não decide posição (ver PLANO_EXPORTADOR_PPTX.md).
            if pptx_template_id and get_settings().PPTX_MUTATE_EXPORT:
                mutated = self._generate_by_mutation(
                    report, user, activities, pptx_template_id,
                    week_number=week_number, year=year, period_label=period_label,
                )
                if mutated is not None:
                    return mutated

            evidence_dossier = self._build_evidence_dossier(activities, user)
            attachment_inventory = self._build_attachment_inventory(activities)

            composed = self.prompt_composer.compose_weekly_prompt(
                user=user,
                evidence_dossier=evidence_dossier,
                week_number=week_number,
                year=year,
                language=lang,
                period_label=period_label,
                attachment_inventory=attachment_inventory,
            )

            def build_plan_prompt(analysis_draft: str) -> tuple[str, str | None]:
                plan = self.prompt_composer.compose_presentation_plan_prompt(
                    user=user,
                    analysis_draft=analysis_draft,
                    attachment_inventory=attachment_inventory,
                    language=lang,
                )
                return plan.user_prompt, plan.system_prompt

            # Layout presente mas sem slides usáveis → trata como sem layout
            # (cai no fluxo de IA em vez de estourar ValueError — QA-030).
            if layout is not None and not (
                isinstance(layout, dict) and layout.get("slides")
            ):
                layout = None

            ai_degraded = False
            if layout is not None and not _layout_uses_ai(layout):
                # Deck 100% manual (nenhum bloco com binding de IA): não há o que
                # a IA preencher → pula a chamada ao LLM. Geração instantânea e
                # fora do gargalo de IA (QA-037).
                structured = {}
                content = "{}"
                logger.info("Weekly manual sem bindings de IA — LLM ignorado")
            else:
                try:
                    content = await self.llm.generate_weekly_content(
                        composed.user_prompt,
                        composed.system_prompt,
                        build_plan_prompt,
                    )
                    structured = self._parse_weekly_content(content)
                    if not structured.get("summary"):
                        raise ValueError("AI response did not contain a summary")
                except Exception as error:
                    logger.warning(
                        "LLM unavailable; generating deterministic weekly | error=%s",
                        error,
                    )
                    ai_degraded = True
                    structured = self._build_fallback_content(activities, lang)
                    content = json.dumps(structured, ensure_ascii=False)

            if layout:
                # Layout do editor WYSIWYG: renderiza exatamente as posições
                # escolhidas pelo usuário (bindings recebem o texto da IA).
                from app.services.pptx_layout import PptxLayoutRenderer

                def _attachment_info(att: Attachment) -> dict:
                    table = None
                    if isinstance(att.kpi_data, dict):
                        table = att.kpi_data.get("table")
                    return {"path": att.file_path, "table": table}

                attachment_map = {
                    att.id: _attachment_info(att)
                    for activity in activities
                    for att in activity.attachments
                }
                # O layout pode referenciar anexos de atividades fora da seleção
                # (ex.: bloco arrastado antes de desmarcar) — busca os que faltam,
                # sempre restrito ao DONO do relatório.
                referenced = {
                    str(el.get("attachment_id"))
                    for sl in layout.get("slides", [])
                    if isinstance(sl, dict)
                    for el in sl.get("elements", [])
                    if isinstance(el, dict) and el.get("attachment_id")
                }
                missing = referenced - set(attachment_map)
                if missing:
                    extra = (
                        self.db.query(Attachment)
                        .join(Activity, Attachment.activity_id == Activity.id)
                        .filter(Attachment.id.in_(missing), Activity.user_id == user.id)
                        .all()
                    )
                    attachment_map.update({att.id: _attachment_info(att) for att in extra})

                # Embute os dados das tabelas no layout salvo: a pré-visualização
                # (inclusive por colegas) renderiza sem depender dos anexos.
                for slide_def in layout.get("slides", []):
                    if not isinstance(slide_def, dict):
                        continue
                    for el in slide_def.get("elements", []):
                        if (
                            isinstance(el, dict)
                            and el.get("type") == "table"
                            and el.get("attachment_id") in attachment_map
                        ):
                            table = attachment_map[el["attachment_id"]].get("table")
                            if table:
                                el["table"] = table

                renderer = PptxLayoutRenderer(
                    Path(get_settings().UPLOAD_DIR) / "reports"
                )
                pptx_path = renderer.generate(
                    report.id, layout, structured, attachment_map
                )
                report.content = {
                    "raw": content,
                    "structured": structured,
                    "ai_degraded": ai_degraded,
                    "layout": layout,
                    "period": {
                        "start": start_date.isoformat() if start_date else None,
                        "end": end_date.isoformat() if end_date else None,
                    },
                    "activity_ids": [activity.id for activity in activities],
                }
                report.pptx_path = pptx_path
                report.prompt_used = composed.full_prompt
                report.ai_summary = str(structured.get("summary", ""))[:1000]
                report.coverage = None
                report.confidence_index = None
                report.quality_score = None
                report.status = WeeklyStatus.COMPLETED
                report.generated_at = datetime.now(UTC)
                for activity in activities:
                    activity.status = ActivityStatus.USED_IN_REPORT

                # Cada montagem gerada ensina o padrão pessoal do usuário
                # (manual pesa mais que rascunho da IA; nunca falha a geração).
                from app.services.style_learning import learn_from_layout
                learn_from_layout(self.db, user.id, layout, source=layout_source)
                # ...e o perfil de conhecimento (KPIs/entidades do usuário).
                from app.services.knowledge_profile import learn_from_activities
                learn_from_activities(self.db, user.id, activities)

                self.db.commit()
                self.db.refresh(report)
                logger.info(
                    "Weekly report generated (layout custom) | report_id=%s | week=%d/%d",
                    report.id, week_number, year,
                )
                return report

            from app.services.pptx_service import PptxService

            pptx_service = PptxService()
            pptx_data = pptx_service.build_data_from_content(
                activities,
                content,
                structured,
                user.name,
                user.department,
                week_number,
                year,
                title=report.title,
                period_label=period_label,
                period_caption="Período" if lang == Language.PT else "Period",
                sector=getattr(user, "sector", QualitySector.CSI),
            )
            unmatched_images = self._attach_activity_images(
                activities, pptx_data, structured
            )

            pptx_path = pptx_service.generate(
                report_id=report.id,
                template_path=None,
                slides_config={},
                data=pptx_data,
                images=unmatched_images,
                week_number=week_number,
                year=year,
                author=user.name,
                language=lang.value,
            )

            report.content = {
                "raw": content,
                "structured": structured,
                "ai_degraded": ai_degraded,
                "period": {
                    "start": start_date.isoformat() if start_date else None,
                    "end": end_date.isoformat() if end_date else None,
                },
                "activity_ids": [activity.id for activity in activities],
            }
            report.pptx_path = pptx_path
            report.prompt_used = composed.full_prompt
            report.ai_summary = str(structured.get("summary", ""))[:1000]
            report.coverage = None
            report.confidence_index = None
            report.quality_score = None
            report.status = WeeklyStatus.COMPLETED
            report.generated_at = datetime.now(UTC)

            for activity in activities:
                activity.status = ActivityStatus.USED_IN_REPORT

            from app.services.knowledge_profile import learn_from_activities
            learn_from_activities(self.db, user.id, activities)

            self.db.commit()
            self.db.refresh(report)
            logger.info("Weekly report generated | report_id=%s | week=%d/%d", report.id, week_number, year)
            return report

        except Exception as e:
            report.status = WeeklyStatus.FAILED
            self.db.commit()
            logger.error("Weekly generation failed | error=%s", str(e))
            raise

    def get_report(self, report_id: str, user_id: str) -> WeeklyReport:
        report = (
            self.db.query(WeeklyReport)
            .filter(WeeklyReport.id == report_id, WeeklyReport.user_id == user_id)
            .first()
        )
        if not report:
            raise NotFoundError("Weekly Report")
        return report

    def list_reports(self, user_id: str, page: int = 1, page_size: int = 20) -> tuple[list[WeeklyReport], int]:
        query = self.db.query(WeeklyReport).filter(WeeklyReport.user_id == user_id)
        total = query.count()
        reports = (
            query.order_by(WeeklyReport.created_at.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        return reports, total

    def _attachment_in_weekly(self, att: Attachment) -> bool:
        if not getattr(att, "include_in_weekly", True):
            return False
        if att.image_usage == ImageUsage.STORE_ONLY:
            return False
        return True

    def _build_attachment_inventory(self, activities: list[Activity]) -> str:
        lines = []
        for index, activity in enumerate(
            sorted(activities, key=lambda item: item.activity_date), start=1
        ):
            for att in activity.attachments:
                if not self._attachment_in_weekly(att):
                    continue
                kpi_hint = ""
                if att.kpi_data and att.kpi_data.get("kpis"):
                    kpi_hint = f" | KPIs: {self._as_evidence_text(att.kpi_data['kpis'][:3])}"
                lines.append(
                    f"- attachment_id={att.id} | activity={index} | type={att.file_type} | "
                    f"file={att.original_filename}{kpi_hint}"
                )
        return "\n".join(lines) if lines else "No attachments included in this weekly."

    def _attach_activity_images(
        self,
        activities: list[Activity],
        pptx_data: dict,
        structured: dict | None = None,
    ) -> list[dict]:
        """Resolve image paths from activity attachments and AI block attachment_ids."""
        ordered = sorted(activities, key=lambda item: item.activity_date)
        att_by_id = {
            att.id: att
            for activity in ordered
            for att in activity.attachments
            if att.file_type == "image" and self._attachment_in_weekly(att)
        }
        images_by_source: dict[int, list[dict]] = {}
        for index, activity in enumerate(ordered, start=1):
            imgs = [
                {
                    "attachment_id": att.id,
                    "path": att.file_path,
                    "caption": att.manual_caption or att.ai_caption or "",
                    "activity": activity.title,
                }
                for att in activity.attachments
                if att.file_type == "image"
                and self._attachment_in_weekly(att)
                and att.file_path
                and Path(att.file_path).exists()
            ]
            if imgs:
                images_by_source[index] = imgs

        if not images_by_source and not att_by_id:
            return []

        used: set[int] = set()
        entries = pptx_data.get("activities")
        if isinstance(entries, list):
            for item in entries:
                if not isinstance(item, dict):
                    continue
                self._resolve_block_images(item, att_by_id)
                source = item.get("source")
                try:
                    source = int(source)
                except (TypeError, ValueError):
                    source = None
                if source not in images_by_source or source in used:
                    source = self._match_activity_by_title(
                        item, ordered, images_by_source, used
                    )
                if source is not None:
                    if not item.get("images"):
                        item["images"] = images_by_source[source]
                    used.add(source)

        if structured and isinstance(structured.get("activities"), list):
            for item, pptx_item in zip(structured["activities"], entries or []):
                if isinstance(item, dict) and isinstance(pptx_item, dict):
                    pptx_item["blocks"] = item.get("blocks") or pptx_item.get("blocks")

        return [
            image
            for index, imgs in images_by_source.items()
            if index not in used
            for image in imgs
        ]

    def _resolve_block_images(self, activity: dict, att_by_id: dict[str, Attachment]) -> None:
        resolved_images = []
        for block in activity.get("blocks") or []:
            if not isinstance(block, dict) or block.get("type") != "image_row":
                continue
            for ref in block.get("images") or []:
                if not isinstance(ref, dict):
                    continue
                att_id = ref.get("attachment_id")
                att = att_by_id.get(att_id) if att_id else None
                if att and Path(att.file_path).exists():
                    resolved_images.append(
                        {
                            "attachment_id": att.id,
                            "path": att.file_path,
                            "caption": ref.get("caption")
                            or att.manual_caption
                            or att.ai_caption
                            or "",
                        }
                    )
            if resolved_images:
                block["images"] = resolved_images
                if not activity.get("images"):
                    activity["images"] = resolved_images[:3]

    def _match_activity_by_title(
        self,
        item: dict,
        ordered: list[Activity],
        images_by_source: dict[int, list[dict]],
        used: set[int],
    ) -> int | None:
        entry_title = str(item.get("title") or "").lower()
        if not entry_title:
            return None
        for index, activity in enumerate(ordered, start=1):
            if index in used or index not in images_by_source:
                continue
            original = activity.title.lower()
            if original[:24] in entry_title or entry_title[:24] in original:
                return index
        return None

    def _build_report_title(
        self,
        start_date: date | None,
        end_date: date | None,
        week_number: int,
        year: int,
        language: Language = Language.PT,
    ) -> str:
        connector = "to" if language == Language.EN else "a"
        week_word = "Week" if language == Language.EN else "Semana"
        if start_date and end_date:
            return (
                f"Weekly Report - {start_date.strftime('%d/%m/%Y')} {connector} "
                f"{end_date.strftime('%d/%m/%Y')}"
            )
        return f"Weekly Report - {week_word} {week_number}/{year}"

    def _build_fallback_content(
        self, activities: list[Activity], language: Language = Language.PT
    ) -> dict:
        ordered = sorted(activities, key=lambda item: item.activity_date)
        activity_items = []
        for index, activity in enumerate(ordered, start=1):
            metadata = activity.metadata_entry
            narrative = (
                activity.description
                or (metadata.technical_summary if metadata else None)
                or activity.title
            )
            activity_items.append(
                {
                    "source": index,
                    "title": activity.title,
                    "date": activity.activity_date.strftime("%d/%m"),
                    "narrative": narrative,
                    "impact": "",
                    "facts": [],
                    "actions": [],
                }
            )

        if language == Language.EN:
            texts = {
                "summary": (
                    f"{len(ordered)} activities were registered in the period. "
                    "The content was organized chronologically based on the provided information."
                ),
                "conclusion": "The registered activities were consolidated for review.",
                "next_step": "Review the content and complete impacts that were not quantified.",
            }
        else:
            texts = {
                "summary": (
                    f"Foram registradas {len(ordered)} atividades no período. "
                    "O conteúdo foi organizado cronologicamente com base nas informações fornecidas."
                ),
                "conclusion": "As atividades registradas foram consolidadas para revisão.",
                "next_step": "Revisar o conteúdo e complementar impactos que não foram quantificados.",
            }

        return {
            "summary": texts["summary"],
            "highlights": [activity.title for activity in ordered[:4]],
            "activities": activity_items,
            "kpis": [],
            "kpi_table": [],
            "conclusions": [texts["conclusion"]],
            "next_steps": [texts["next_step"]],
            "archetype": "operacional",
            "narrative": "\n".join(
                f"{item['date']} — {item['title']}: {item['narrative']}"
                for item in activity_items
            ),
            "presentation_plan": {
                "layout_profile": "executive",
                "sidebar": [],
                "global_blocks": [],
            },
        }

    async def _ensure_attachments_analyzed(self, activities: list[Activity]) -> None:
        """Re-run AI analysis for attachments that were skipped or failed in background."""
        file_service = FileService(self.db, self.llm)
        for activity in activities:
            analyze_images = activity_requests_image_analysis(activity)
            for attachment in activity.attachments:
                if attachment.file_type == "image" and not analyze_images:
                    continue
                analysis = attachment.ai_analysis or {}
                needs_analysis = not analysis or analysis.get("status") == "pending_ai"
                if not needs_analysis:
                    continue
                try:
                    await file_service.process_attachment(attachment)
                except Exception as error:
                    logger.warning(
                        "On-demand attachment analysis failed | attachment_id=%s | error=%s",
                        attachment.id,
                        error,
                    )

    def _build_evidence_dossier(self, activities: list[Activity], user: User) -> str:
        """Full per-activity dossier with sector context and attachment references."""
        sector = getattr(user, "sector", QualitySector.CSI)
        blocks = [f"Author sector: {sector.value}"]
        for index, activity in enumerate(
            sorted(activities, key=lambda item: item.activity_date), start=1
        ):
            directives = parse_activity_directives(activity.description)
            registration_text = (
                directives.clean_description or "N/A"
            )
            lines = [
                f"### Activity {index}: {activity.title}",
                f"Date: {activity.activity_date.strftime('%A %d/%m/%Y %H:%M')}",
                f"Registration text: {registration_text}",
                f"analyze_images_requested: {directives.analyze_images}",
            ]

            meta = activity.metadata_entry
            if meta:
                meta_parts = []
                for label, value in (
                    ("project", meta.project),
                    ("supplier", meta.supplier),
                    ("line", meta.line),
                    ("process", meta.process),
                    ("product", meta.product),
                    ("category", meta.category),
                    ("activity_type", meta.activity_type),
                    ("defect_type", meta.defect_type),
                ):
                    if value:
                        meta_parts.append(f"{label}={value}")
                if meta.related_kpis:
                    meta_parts.append(f"related_kpis={', '.join(map(str, meta.related_kpis))}")
                if meta_parts:
                    lines.append(f"Extracted metadata: {'; '.join(meta_parts)}")
                if meta.technical_summary:
                    lines.append(f"Technical summary: {meta.technical_summary}")

            attachments = [
                att for att in activity.attachments if self._attachment_in_weekly(att)
            ]
            for att in attachments:
                lines.append(
                    self._describe_attachment(
                        att, analyze_images=directives.analyze_images
                    )
                )

            if not attachments:
                lines.append("Attachments: none.")

            blocks.append("\n".join(lines))

        return "\n\n".join(blocks)

    def _spreadsheet_markdown_preview(self, kpi_data: dict) -> str | None:
        preview = kpi_data.get("preview")
        if not preview:
            return None
        if isinstance(preview, list) and preview and isinstance(preview[0], dict):
            headers = list(preview[0].keys())[:8]
            rows = preview[:12]
            header_line = "| " + " | ".join(headers) + " |"
            sep_line = "| " + " | ".join("---" for _ in headers) + " |"
            body = []
            for row in rows:
                body.append("| " + " | ".join(str(row.get(h, ""))[:24] for h in headers) + " |")
            return "\n".join([header_line, sep_line, *body])
        return str(preview)[:1500]

    def _describe_attachment(
        self, att: Attachment, *, analyze_images: bool = False
    ) -> str:
        header = (
            f"Attachment [id={att.id}] [{att.file_type}] {att.original_filename}:"
        )
        parts: list[str] = []

        if att.manual_caption:
            parts.append(f"user caption: {att.manual_caption}")

        analysis = att.ai_analysis or {}
        if analysis.get("status") == "pending_ai":
            analysis = {}

        if att.file_type == "image":
            if not analyze_images:
                parts.append("visual reference only")
            else:
                caption = analysis.get("caption") or att.ai_caption
                if caption:
                    parts.append(f"image content: {self._as_evidence_text(caption)}")
                if analysis.get("visible_text"):
                    parts.append(
                        f"visible text in image: {self._as_evidence_text(analysis['visible_text'])}"
                    )
                if analysis.get("observations"):
                    parts.append(
                        f"observations: {self._as_evidence_text(analysis['observations'])}"
                    )
                if analysis.get("possible_impact"):
                    parts.append(
                        f"possible impact: {self._as_evidence_text(analysis['possible_impact'])}"
                    )
                if analysis.get("device_info"):
                    parts.append(
                        f"device info: {self._as_evidence_text(analysis['device_info'])}"
                    )
        elif att.file_type == "spreadsheet":
            kpi_data = att.kpi_data or {}
            if kpi_data.get("summary") or analysis.get("summary"):
                parts.append(
                    f"data summary: {self._as_evidence_text(kpi_data.get('summary') or analysis.get('summary'))}"
                )
            if kpi_data.get("kpis") or analysis.get("kpis"):
                parts.append(
                    f"KPIs found: {self._as_evidence_text(kpi_data.get('kpis') or analysis.get('kpis'))}"
                )
            if analysis.get("measurements") or kpi_data.get("measurements"):
                parts.append(
                    f"measurements: {self._as_evidence_text(analysis.get('measurements') or kpi_data.get('measurements'))}"
                )
            if analysis.get("time_series") or kpi_data.get("time_series"):
                parts.append(
                    f"time series: {self._as_evidence_text(analysis.get('time_series') or kpi_data.get('time_series'))}"
                )
            if kpi_data.get("trends") or analysis.get("trends"):
                parts.append(
                    f"trends: {self._as_evidence_text(kpi_data.get('trends') or analysis.get('trends'))}"
                )
            if kpi_data.get("anomalies") or analysis.get("anomalies"):
                parts.append(
                    f"anomalies: {self._as_evidence_text(kpi_data.get('anomalies') or analysis.get('anomalies'))}"
                )
            table_preview = self._spreadsheet_markdown_preview(kpi_data)
            if table_preview:
                parts.append(f"tabular preview:\n{table_preview}")
        else:
            if analysis.get("summary"):
                parts.append(f"document summary: {self._as_evidence_text(analysis['summary'])}")
            if analysis.get("relevant_facts"):
                parts.append(f"relevant facts: {self._as_evidence_text(analysis['relevant_facts'])}")
            if analysis.get("possible_impacts"):
                parts.append(f"possible impacts: {self._as_evidence_text(analysis['possible_impacts'])}")
            if analysis.get("kpis"):
                parts.append(f"KPIs mentioned: {self._as_evidence_text(analysis['kpis'])}")

        if not parts and analysis.get("raw"):
            parts.append(f"raw analysis: {str(analysis['raw'])[:1200]}")
        if not parts:
            parts.append("stored, but no analysis is available for this file.")

        return header + "\n  - " + "\n  - ".join(parts)

    def _as_evidence_text(self, value) -> str:
        if isinstance(value, list):
            return "; ".join(self._as_evidence_text(item) for item in value if item)
        if isinstance(value, dict):
            return "; ".join(
                f"{key}: {self._as_evidence_text(item)}" for key, item in value.items() if item
            )
        return str(value).strip()

    def _calculate_coverage(self, activities: list[Activity], template) -> CoverageMetrics:
        images = sum(
            1 for a in activities for att in a.attachments
            if att.file_type == "image" and self._attachment_in_weekly(att)
        )
        files = sum(
            1 for a in activities for att in a.attachments if self._attachment_in_weekly(att)
        )
        kpis = sum(
            len(a.metadata_entry.related_kpis) if a.metadata_entry and a.metadata_entry.related_kpis else 0
            for a in activities
        )

        slides_config = template.slides_config or {}
        total_slides = len(slides_config.get("slides", [])) or 5
        filled_slides = min(len(activities), total_slides)

        quality = min(100.0, (len(activities) * 10 + images * 5 + kpis * 3) / max(total_slides, 1))

        return CoverageMetrics(
            activities_registered=len(activities),
            activities_used=len(activities),
            images_used=images,
            files_used=files,
            kpis_identified=kpis,
            slides_filled=filled_slides,
            missing_required_fields=[],
            quality_score=round(quality, 1),
        )

    def _calculate_confidence(self, activities: list[Activity], template) -> list:
        from app.schemas.weekly import ConfidenceSlide

        slides_config = template.slides_config or {}
        slides = slides_config.get("slides", [])
        if not slides:
            slides = [{"title": "Summary", "number": 1}, {"title": "Activities", "number": 2}]

        confidence_slides = []
        for slide in slides:
            conf = min(95.0, 50.0 + len(activities) * 5)
            missing = []
            if len(activities) < 3:
                missing.append("Insufficient activities for comprehensive coverage")
            confidence_slides.append(
                ConfidenceSlide(
                    slide_number=slide.get("number", 0),
                    slide_title=slide.get("title", "Untitled"),
                    confidence=round(conf, 1),
                    missing_evidence=missing,
                )
            )
        return confidence_slides

    def _parse_weekly_content(self, content: str) -> dict:
        try:
            start = content.find("{")
            end = content.rfind("}") + 1
            if start >= 0 and end > start:
                raw = json.loads(content[start:end])
                validated = parse_weekly_content(raw)
                legacy = weekly_content_to_legacy_dict(validated)
                return sanitize_weekly_content(legacy)
        except (json.JSONDecodeError, ValueError) as error:
            logger.warning("Weekly content validation failed: %s", error)
        try:
            start = content.find("{")
            end = content.rfind("}") + 1
            if start >= 0 and end > start:
                return json.loads(content[start:end])
        except (json.JSONDecodeError, ValueError):
            pass
        return {"narrative": content}
