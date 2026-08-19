"""Extração determinística de tabelas de planilhas (xlsx/xls/csv).

Roda de forma SÍNCRONA no salvamento do anexo — sem IA, sem rede.
Eficiência: openpyxl em read_only, primeira planilha com dados,
limitada a MAX_ROWS × MAX_COLS. O resultado vai para Attachment.kpi_data
como {"table": {...}} e é usado nos bloquinhos do editor e no PPTX.
"""

import csv
import io
import logging
from datetime import date, datetime, time
from pathlib import Path

logger = logging.getLogger(__name__)

MAX_ROWS = 30
MAX_COLS = 12


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y") if value.time() == time.min else value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def _trim(rows: list[list[str]]) -> dict | None:
    """Remove linhas/colunas totalmente vazias das bordas e monta o payload."""
    rows = [r for r in rows if any(c for c in r)]
    if not rows:
        return None
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    # corta colunas vazias à direita
    while width > 0 and all(not r[width - 1] for r in rows):
        width -= 1
    if width == 0:
        return None
    truncated = len(rows) > MAX_ROWS or width > MAX_COLS
    keep = min(width, MAX_COLS)
    rows = [r[:keep] for r in rows[:MAX_ROWS]]
    return {
        "columns": rows[0],
        "rows": rows[1:],
        "n_rows": len(rows) - 1,
        "n_cols": min(width, MAX_COLS),
        "truncated": truncated,
    }


def _extract_xlsx(content: bytes) -> dict | None:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows: list[list[str]] = []
            for row in sheet.iter_rows(max_row=MAX_ROWS * 3, max_col=MAX_COLS, values_only=True):
                rows.append([_cell_to_str(cell) for cell in row])
                if len([r for r in rows if any(r)]) >= MAX_ROWS + 1:
                    break
            table = _trim(rows)
            if table:
                table["sheet"] = sheet.title
                return table
        return None
    finally:
        workbook.close()


def _extract_xls(content: bytes) -> dict | None:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    for sheet in book.sheets():
        rows = [
            [_cell_to_str(sheet.cell_value(r, c)) for c in range(min(sheet.ncols, MAX_COLS))]
            for r in range(min(sheet.nrows, MAX_ROWS * 3))
        ]
        table = _trim(rows)
        if table:
            table["sheet"] = sheet.name
            return table
    return None


def _extract_csv(content: bytes) -> dict | None:
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows = [[cell.strip() for cell in row[:MAX_COLS]] for _, row in zip(range(MAX_ROWS * 3), reader)]
    return _trim(rows)


def extract_table(content: bytes, filename: str) -> dict | None:
    """Tabela extraída da planilha ou None (arquivo vazio/ilegível nunca falha o upload)."""
    ext = Path(filename).suffix.lower()
    try:
        if ext == ".xlsx":
            return _extract_xlsx(content)
        if ext == ".xls":
            return _extract_xls(content)
        if ext == ".csv":
            return _extract_csv(content)
        return None
    except Exception as error:
        logger.warning("Falha ao extrair tabela | file=%s | err=%s", filename, error)
        return None


# ── Leitura COMPLETA para cálculo (o extract_table acima é preview 30×12) ────

MAX_ROWS_FULL = 20000
MAX_COLS_FULL = 60



def _trim_full(rows: list[list[str]]) -> dict | None:
    """Corta colunas vazias à direita (cabeçalho vazio) e devolve o payload."""
    if len(rows) < 2:
        return None
    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]
    header = rows[0]
    # largura útil = última coluna com CABEÇALHO preenchido
    while width > 0 and not str(header[width - 1]).strip():
        width -= 1
    if width == 0:
        return None
    rows = [r[:width] for r in rows]
    return {"columns": rows[0], "rows": rows[1:]}


def extract_full_table(content: bytes, filename: str) -> dict | None:
    """Planilha inteira (até MAX_ROWS_FULL linhas) para o motor de análise.

    Diferente de `extract_table` (preview truncado para a UI/PPT), aqui
    precisamos de TODAS as linhas — os cálculos agregam a base completa.
    Devolve {"columns": [...], "rows": [[...], ...]} ou None.
    """
    ext = Path(filename).suffix.lower()
    try:
        if ext == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            for sheet in workbook.worksheets:
                rows: list[list[str]] = []
                for row in sheet.iter_rows(max_row=MAX_ROWS_FULL, max_col=MAX_COLS_FULL,
                                           values_only=True):
                    rows.append([_cell_to_str(v) for v in row])
                rows = [r for r in rows if any(c for c in r)]
                if len(rows) >= 2:
                    workbook.close()
                    return _trim_full(rows)
            workbook.close()
            return None
        if ext == ".csv":
            text = content.decode("utf-8-sig", errors="replace")
            sample = text[:4096]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
            except csv.Error:
                dialect = csv.excel
            rows = [
                [_cell_to_str(c) for c in row]
                for row in csv.reader(io.StringIO(text), dialect)
            ]
            rows = [r for r in rows if any(c for c in r)][:MAX_ROWS_FULL]
            return _trim_full(rows)
        if ext == ".xls":
            import xlrd

            book = xlrd.open_workbook(file_contents=content)
            sheet = book.sheet_by_index(0)
            rows = [
                [_cell_to_str(sheet.cell_value(r, c)) for c in range(min(sheet.ncols, MAX_COLS_FULL))]
                for r in range(min(sheet.nrows, MAX_ROWS_FULL))
            ]
            rows = [r for r in rows if any(c for c in r)]
            return _trim_full(rows)
        return None
    except Exception as error:
        logger.warning("Falha na leitura completa | file=%s | err=%s", filename, error)
        return None
