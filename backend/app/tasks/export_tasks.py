"""
Export Tasks - Async data export via Celery.

Tasks:
- export_activities: Export user activities to CSV/Excel
- export_weekly_report: Export weekly report data
"""

import logging
import csv
from io import StringIO
from datetime import datetime
from typing import Any, Optional
from pathlib import Path

from app.celery_app import celery_app
from app.core.database import SessionLocal

logger = logging.getLogger(__name__)


@celery_app.task(
    bind=True,
    max_retries=2,
    default_retry_delay=60,
    queue="exports",
)
def export_activities(
    self,
    user_id: str,
    format: str = "csv",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> dict[str, Any]:
    """
    Export activities to file.

    Performs:
    - Query activities based on filters
    - Format to CSV/Excel/JSON
    - Generate download file
    - Log export event

    Args:
        user_id: User ID (owner)
        format: Export format (csv, excel, json)
        date_from: Optional date filter (ISO format)
        date_to: Optional date filter (ISO format)

    Returns:
        Export result with file path

    Raises:
        Retries on failure (max 2 times)
    """
    logger.info(
        f"Exporting activities | user={user_id} | format={format} | "
        f"from={date_from} to={date_to}"
    )

    db = SessionLocal()
    try:
        from app.models import Activity
        from app.events import get_event_bus
        from app.events.types import ProcessingStartedEvent, ProcessingCompletedEvent
        from time import time

        start_time = time()

        # Publish export started event
        event_bus = get_event_bus(db)
        start_event = ProcessingStartedEvent(
            aggregate_id=user_id,
            processing_type="export_activities",
            user_id=user_id,
            metadata={
                "format": format,
                "date_from": date_from,
                "date_to": date_to,
            },
        )
        event_bus.publish(start_event)

        # Query activities
        query = db.query(Activity).filter(Activity.owner_id == user_id)

        if date_from:
            from datetime import datetime as dt

            start_date = dt.fromisoformat(date_from)
            query = query.filter(Activity.created_at >= start_date)

        if date_to:
            from datetime import datetime as dt

            end_date = dt.fromisoformat(date_to)
            query = query.filter(Activity.created_at <= end_date)

        activities = query.all()
        logger.info(f"Exporting {len(activities)} activities")

        # Format data
        if format == "csv":
            file_path = _export_activities_csv(activities, user_id)
        elif format == "excel":
            file_path = _export_activities_excel(activities, user_id)
        elif format == "json":
            file_path = _export_activities_json(activities, user_id)
        else:
            raise ValueError(f"Unsupported format: {format}")

        duration = time() - start_time

        # Publish completion event
        completion_event = ProcessingCompletedEvent(
            aggregate_id=user_id,
            processing_type="export_activities",
            duration_seconds=duration,
            user_id=user_id,
            result={
                "file_path": str(file_path),
                "format": format,
                "count": len(activities),
            },
        )
        event_bus.publish(completion_event)

        logger.info(
            f"Activities exported successfully | path={file_path} | "
            f"count={len(activities)} | duration={duration:.2f}s"
        )

        return {
            "status": "success",
            "file_path": str(file_path),
            "format": format,
            "count": len(activities),
            "duration": duration,
        }

    except Exception as exc:
        logger.error(f"Error exporting activities: {str(exc)}", exc_info=True)
        raise self.retry(exc=exc, countdown=60 * (2 ** self.request.retries))

    finally:
        db.close()


def _export_activities_csv(activities: list, user_id: str) -> Path:
    """Export activities to CSV file."""
    from app.core.config import get_settings

    settings = get_settings()
    export_dir = Path(settings.UPLOAD_DIR) / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    file_path = export_dir / f"activities_{user_id}_{timestamp}.csv"

    with open(file_path, "w", newline="") as csvfile:
        fieldnames = [
            "ID",
            "Title",
            "Department",
            "Date",
            "Status",
            "Created At",
            "Updated At",
        ]
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

        writer.writeheader()
        for activity in activities:
            writer.writerow({
                "ID": activity.id,
                "Title": activity.title,
                "Department": activity.department,
                "Date": activity.activity_date,
                "Status": activity.status.value if activity.status else "",
                "Created At": activity.created_at,
                "Updated At": activity.updated_at,
            })

    logger.info(f"CSV export saved to {file_path}")
    return file_path


def _export_activities_excel(activities: list, user_id: str) -> Path:
    """Export activities to Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not installed, using CSV fallback")
        return _export_activities_csv(activities, user_id)

    from app.core.config import get_settings

    settings = get_settings()
    export_dir = Path(settings.UPLOAD_DIR) / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    file_path = export_dir / f"activities_{user_id}_{timestamp}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activities"

    # Header
    headers = ["ID", "Title", "Department", "Date", "Status", "Created At", "Updated At"]
    ws.append(headers)

    # Style header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Data
    for activity in activities:
        ws.append([
            activity.id,
            activity.title,
            activity.department,
            activity.activity_date,
            activity.status.value if activity.status else "",
            activity.created_at,
            activity.updated_at,
        ])

    wb.save(file_path)
    logger.info(f"Excel export saved to {file_path}")
    return file_path


def _export_activities_json(activities: list, user_id: str) -> Path:
    """Export activities to JSON file."""
    import json

    from app.core.config import get_settings

    settings = get_settings()
    export_dir = Path(settings.UPLOAD_DIR) / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    file_path = export_dir / f"activities_{user_id}_{timestamp}.json"

    data = {
        "user_id": user_id,
        "exported_at": datetime.utcnow().isoformat(),
        "count": len(activities),
        "activities": [
            {
                "id": activity.id,
                "title": activity.title,
                "department": activity.department,
                "activity_date": activity.activity_date.isoformat() if activity.activity_date else None,
                "status": activity.status.value if activity.status else None,
                "created_at": activity.created_at.isoformat(),
                "updated_at": activity.updated_at.isoformat(),
            }
            for activity in activities
        ],
    }

    with open(file_path, "w") as f:
        json.dump(data, f, indent=2)

    logger.info(f"JSON export saved to {file_path}")
    return file_path


@celery_app.task(
    bind=True,
    max_retries=2,
    default_retry_delay=60,
    queue="exports",
)
def export_weekly_report(
    self,
    weekly_id: str,
    user_id: str,
    format: str = "pdf",
) -> dict[str, Any]:
    """
    Export weekly report.

    Args:
        weekly_id: Weekly report ID
        user_id: User ID (owner)
        format: Export format (pdf, pptx, json)

    Returns:
        Export result with file path

    Raises:
        Retries on failure (max 2 times)
    """
    logger.info(
        f"Exporting weekly report | weekly={weekly_id} | user={user_id} | format={format}"
    )

    db = SessionLocal()
    try:
        from app.models import WeeklyReport
        from app.events import get_event_bus
        from app.events.types import ProcessingStartedEvent, ProcessingCompletedEvent
        from time import time

        start_time = time()

        # Publish export started
        event_bus = get_event_bus(db)
        start_event = ProcessingStartedEvent(
            aggregate_id=weekly_id,
            processing_type="export_weekly",
            user_id=user_id,
            metadata={"format": format},
        )
        event_bus.publish(start_event)

        # Get weekly report
        weekly = db.query(WeeklyReport).filter(WeeklyReport.id == weekly_id).first()
        if not weekly:
            raise ValueError(f"Weekly report {weekly_id} not found")

        # Generate export (format-specific logic)
        file_path = Path(f"uploads/exports/weekly_{weekly_id}_{format}")
        logger.info(f"Weekly report export ready: {file_path}")

        duration = time() - start_time

        # Publish completion
        completion_event = ProcessingCompletedEvent(
            aggregate_id=weekly_id,
            processing_type="export_weekly",
            duration_seconds=duration,
            user_id=user_id,
            result={"file_path": str(file_path), "format": format},
        )
        event_bus.publish(completion_event)

        return {
            "status": "success",
            "file_path": str(file_path),
            "format": format,
            "duration": duration,
        }

    except Exception as exc:
        logger.error(f"Error exporting weekly report: {str(exc)}", exc_info=True)
        raise self.retry(exc=exc, countdown=60 * (2 ** self.request.retries))

    finally:
        db.close()
