"""Diagnóstico do DRM no servidor Windows — responde, com evidência, COMO o
QWI vai conseguir ler os arquivos protegidos.

Rode ISTO no PC servidor, com um arquivo protegido de verdade em mãos, ANTES
de qualquer implementação. O resultado diz qual dos três caminhos existe:

  1. Processo autorizado — o Python já lê em claro. É o melhor caso: nada
     precisa ser implementado, e o deploy continua como está.
  2. SDK/CLI do fornecedor — não é detectável por aqui; pergunte ao time de
     segurança. Seria um adaptador pequeno, sem Office.
  3. Ponte via Office COM — funciona, mas exige sessão interativa logada e
     torna o processamento serial. É o caminho mais caro.

Uso:

    python scripts\\diagnostico_drm.py C:\\caminho\\planilha_protegida.xlsx
    python scripts\\diagnostico_drm.py arq1.xlsx arq2.pptx arq3.pdf

O script é somente-leitura sobre os seus arquivos: quando testa o Office, a
cópia limpa vai para uma pasta temporária e é apagada em seguida.
"""
from __future__ import annotations

import os
import platform
import shutil
import sys
import tempfile
from pathlib import Path

# Assinaturas de início de arquivo. Um .xlsx/.pptx/.docx é um ZIP ("PK"); se
# vier um cabeçalho OLE, é quase certo que o conteúdo está envelopado pelo DRM
# (ou é um formato antigo .xls/.ppt/.doc legítimo — o teste de leitura
# desempata).
ASSINATURAS = {
    b"PK\x03\x04": "ZIP (Office moderno: xlsx/pptx/docx em claro)",
    b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1": "OLE/CFBF (Office 97-2003 OU envelope de DRM)",
    b"%PDF": "PDF em claro",
    b"\xff\xd8\xff": "JPEG em claro",
    b"\x89PNG": "PNG em claro",
}

# Aplicação do Office capaz de abrir cada extensão. PDF e imagem não têm —
# é justamente por isso que a ponte por Office não cobre esses dois.
APP_POR_EXTENSAO = {
    ".xlsx": "Excel", ".xlsm": "Excel", ".xls": "Excel", ".csv": "Excel",
    ".pptx": "PowerPoint", ".ppt": "PowerPoint",
    ".docx": "Word", ".doc": "Word",
}

LARGURA = 78


def titulo(texto: str) -> None:
    print()
    print("─" * LARGURA)
    print(texto)
    print("─" * LARGURA)


def item(rotulo: str, valor: object) -> None:
    print(f"  {rotulo:<38} {valor}")


# ── 1. Ambiente ─────────────────────────────────────────────────────────────

def sessao_interativa() -> tuple[bool, str]:
    """A automação do Office só funciona com desktop (não em Session 0).

    Session 0 é onde ficam os serviços do Windows e as tarefas agendadas
    marcadas como "executar estando o usuário conectado ou não" — que é
    exatamente como o DEPLOY_WINDOWS.md registra o backend hoje.
    """
    if platform.system() != "Windows":
        return False, "não é Windows"
    try:
        import ctypes

        sessao = ctypes.c_ulong()
        pid = ctypes.windll.kernel32.GetCurrentProcessId()
        ok = ctypes.windll.kernel32.ProcessIdToSessionId(pid, ctypes.byref(sessao))
        if not ok:
            return False, "não foi possível ler a sessão"
        numero = sessao.value
        if numero == 0:
            return False, "Session 0 (serviço/sem desktop) — Office COM NÃO funciona aqui"
        return True, f"Session {numero} (interativa)"
    except Exception as erro:  # noqa: BLE001
        return False, f"falha ao consultar: {erro}"


def diagnosticar_ambiente() -> None:
    titulo("1. AMBIENTE")
    item("Sistema", f"{platform.system()} {platform.release()}")
    item("Python", sys.version.split()[0])
    item("Executável", sys.executable)
    item("Usuário", os.environ.get("USERNAME") or os.environ.get("USER") or "?")

    interativa, detalhe = sessao_interativa()
    item("Sessão", detalhe)
    if platform.system() == "Windows" and not interativa:
        print()
        print("  ATENÇÃO: sem sessão interativa, o caminho por Office está fora.")
        print("  Se o teste de leitura direta (item 3) falhar, só restam as")
        print("  opções 1 (autorizar o processo) ou 2 (SDK do fornecedor).")

    try:
        import win32com.client  # noqa: F401
        item("pywin32", "instalado")
    except ImportError:
        item("pywin32", "AUSENTE (pip install pywin32) — necessário só para a opção 3")
    except Exception as erro:  # noqa: BLE001
        item("pywin32", f"erro ao importar: {erro}")


def diagnosticar_agente() -> None:
    """Mostra processos que aparentam ser o agente de DRM.

    Serve para confirmar que o agente está rodando NESTA sessão — um agente
    ativo na sessão do administrador não ajuda o backend rodando em outra.
    """
    titulo("2. AGENTE DE DRM (processos suspeitos nesta máquina)")
    if platform.system() != "Windows":
        item("—", "pulado (só faz sentido no servidor Windows)")
        return
    pistas = ("nasca", "softcamp", "drm", "docsecu", "fasoo", "markany", "sgate")
    try:
        import subprocess

        saida = subprocess.run(
            ["tasklist", "/fo", "csv", "/nh"],
            capture_output=True, text=True, timeout=30,
        ).stdout
    except Exception as erro:  # noqa: BLE001
        item("—", f"não foi possível listar processos: {erro}")
        return

    achados = [
        linha.split(",")[0].strip('"')
        for linha in saida.splitlines()
        if any(pista in linha.lower() for pista in pistas)
    ]
    if achados:
        for nome in sorted(set(achados)):
            item("processo", nome)
        print()
        print("  Leve estes nomes para o time de segurança: são eles que dizem")
        print("  qual é o produto e se existe SDK de servidor (opção 2).")
    else:
        item("—", "nenhum processo reconhecido pelas pistas conhecidas")
        print("  (Não conclua que não há agente: o nome pode ser outro.)")


# ── 3. Leitura direta: o processo já está autorizado? ───────────────────────

def ler_cabecalho(caminho: Path) -> str:
    with caminho.open("rb") as arquivo:
        inicio = arquivo.read(8)
    for assinatura, descricao in ASSINATURAS.items():
        if inicio.startswith(assinatura):
            return f"{inicio[:4].hex(' ')} → {descricao}"
    return f"{inicio.hex(' ')} → desconhecido (provável envelope de DRM)"


def tentar_parse_direto(caminho: Path) -> tuple[bool, str]:
    """Tenta ler com as MESMAS bibliotecas que o QWI usa.

    Se isto der certo, o processo já enxerga o conteúdo em claro: é a opção 1,
    e não há nada a implementar.
    """
    extensao = caminho.suffix.lower()
    try:
        if extensao in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook

            livro = load_workbook(caminho, read_only=True, data_only=True)
            abas = list(livro.sheetnames)
            aba = livro[abas[0]]
            primeira = next(aba.iter_rows(max_row=1, values_only=True), ())
            livro.close()   # depois de ler: em read_only o workbook fechado não responde mais
            return True, f"openpyxl leu {len(abas)} aba(s); 1ª linha: {primeira!r:.60}"
        if extensao in (".pptx", ".ppt"):
            from pptx import Presentation

            apresentacao = Presentation(str(caminho))
            return True, f"python-pptx leu {len(apresentacao.slides)} slide(s)"
        if extensao == ".xls":
            import xlrd

            livro = xlrd.open_workbook(str(caminho))
            return True, f"xlrd leu {livro.nsheets} aba(s)"
        if extensao == ".pdf":
            with caminho.open("rb") as arquivo:
                if arquivo.read(4) == b"%PDF":
                    return True, "cabeçalho %PDF presente (não envelopado)"
            return False, "não começa com %PDF"
        if extensao in (".png", ".jpg", ".jpeg"):
            from PIL import Image

            with Image.open(caminho) as imagem:
                return True, f"Pillow leu {imagem.format} {imagem.size}"
        return False, f"extensão {extensao} sem teste direto neste script"
    except Exception as erro:  # noqa: BLE001
        return False, f"{type(erro).__name__}: {str(erro)[:90]}"


# ── 4. Ponte por Office COM ─────────────────────────────────────────────────

def tentar_office(caminho: Path) -> tuple[bool, str]:
    """Abre pelo Office e salva uma cópia limpa numa pasta temporária.

    É o teste da opção 3. A cópia é apagada no fim — seus arquivos não são
    tocados.
    """
    extensao = caminho.suffix.lower()
    app_nome = APP_POR_EXTENSAO.get(extensao)
    if not app_nome:
        return False, f"nenhuma aplicação do Office abre {extensao} (PDF/imagem não passam por aqui)"
    if platform.system() != "Windows":
        return False, "pulado (só roda no servidor Windows)"

    try:
        import pythoncom
        import win32com.client
    except ImportError:
        return False, "pywin32 não instalado"

    destino = Path(tempfile.mkdtemp(prefix="qwi_drm_")) / f"limpo{extensao}"
    pythoncom.CoInitialize()
    aplicacao = None
    try:
        if app_nome == "Excel":
            aplicacao = win32com.client.DispatchEx("Excel.Application")
            aplicacao.Visible = False
            aplicacao.DisplayAlerts = False
            documento = aplicacao.Workbooks.Open(str(caminho), ReadOnly=True)
            documento.SaveAs(str(destino), FileFormat=51)  # 51 = xlsx
            documento.Close(SaveChanges=False)
        elif app_nome == "PowerPoint":
            aplicacao = win32com.client.DispatchEx("PowerPoint.Application")
            documento = aplicacao.Presentations.Open(
                str(caminho), ReadOnly=True, WithWindow=False
            )
            documento.SaveAs(str(destino), 24)  # 24 = pptx
            documento.Close()
        else:  # Word
            aplicacao = win32com.client.DispatchEx("Word.Application")
            aplicacao.Visible = False
            documento = aplicacao.Documents.Open(str(caminho), ReadOnly=True)
            documento.SaveAs2(str(destino), FileFormat=16)  # 16 = docx
            documento.Close(SaveChanges=False)

        if not destino.exists():
            return False, "o Office não gerou a cópia"
        tamanho = destino.stat().st_size
        cabecalho = ler_cabecalho(destino)
        return True, f"cópia limpa com {tamanho} bytes — {cabecalho}"
    except Exception as erro:  # noqa: BLE001
        return False, f"{type(erro).__name__}: {str(erro)[:120]}"
    finally:
        # Sem isto sobra EXCEL.EXE/POWERPNT.EXE órfão consumindo a máquina.
        try:
            if aplicacao is not None:
                aplicacao.Quit()
        except Exception:  # noqa: BLE001
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:  # noqa: BLE001
            pass
        shutil.rmtree(destino.parent, ignore_errors=True)


def diagnosticar_arquivo(caminho: Path) -> dict:
    titulo(f"ARQUIVO: {caminho.name}")
    if not caminho.exists():
        item("erro", "arquivo não encontrado")
        return {"arquivo": caminho.name, "direto": False, "office": False}

    item("Caminho", caminho)
    item("Tamanho", f"{caminho.stat().st_size} bytes")
    item("Cabeçalho", ler_cabecalho(caminho))

    print()
    direto_ok, direto_msg = tentar_parse_direto(caminho)
    item("Leitura direta (Python)", "OK" if direto_ok else "FALHOU")
    item("  detalhe", direto_msg)

    print()
    office_ok, office_msg = tentar_office(caminho)
    item("Leitura via Office (COM)", "OK" if office_ok else "FALHOU")
    item("  detalhe", office_msg)

    return {"arquivo": caminho.name, "direto": direto_ok, "office": office_ok}


def veredito(resultados: list[dict]) -> None:
    titulo("VEREDITO")
    if not resultados:
        print("  Nenhum arquivo testado. Passe ao menos um arquivo protegido.")
        return

    todos_diretos = all(r["direto"] for r in resultados)
    algum_direto = any(r["direto"] for r in resultados)
    algum_office = any(r["office"] for r in resultados)
    sem_saida = [r["arquivo"] for r in resultados if not r["direto"] and not r["office"]]

    for r in resultados:
        estado = (
            "já legível pelo Python" if r["direto"]
            else "só pelo Office" if r["office"]
            else "NENHUM método funcionou"
        )
        item(r["arquivo"], estado)

    print()
    if todos_diretos:
        print("  OPÇÃO 1 — o processo já está autorizado no DRM.")
        print("  Não há nada a implementar: o upload de anexos e templates já")
        print("  funciona com esses arquivos. Confirme com um arquivo protegido")
        print("  de verdade (não uma cópia já liberada) antes de concluir.")
    elif algum_office:
        print("  OPÇÃO 3 — a ponte por Office funciona nesta máquina.")
        print("  Antes de implementar, pergunte ao time de segurança sobre a")
        print("  OPÇÃO 1 (autorizar o processo do backend) e a OPÇÃO 2 (SDK de")
        print("  servidor): as duas são mais simples e mais estáveis que esta.")
        print("  Se for por aqui, o backend passa a exigir sessão interativa —")
        print("  o registro atual no Agendador de Tarefas precisa mudar.")
    elif algum_direto:
        print("  RESULTADO MISTO — alguns arquivos leem direto, outros não.")
        print("  Ou os que leram não estavam protegidos, ou a autorização do")
        print("  processo cobre só parte dos formatos. Repita o teste usando,")
        print("  para CADA formato, um arquivo comprovadamente protegido.")
    else:
        print("  NENHUM caminho funcionou nesta máquina.")
        print("  Verifique se o agente de DRM está ativo NESTA sessão e se o")
        print("  usuário logado tem permissão sobre estes arquivos. Se estiver")
        print("  tudo certo e ainda assim falhar, a saída é a OPÇÃO 2 (SDK do")
        print("  fornecedor) — leve os nomes de processo do item 2 à segurança.")

    if sem_saida and (todos_diretos or algum_office or algum_direto):
        print()
        print(f"  Sem nenhum caminho: {', '.join(sem_saida)}")

    print()
    print("  Lembrete: PDF e imagem não abrem pelo Office. Se esses formatos")
    print("  também chegam protegidos, só as opções 1 ou 2 os atendem.")


def main() -> int:
    print("=" * LARGURA)
    print("QWI — diagnóstico de arquivos protegidos por DRM".center(LARGURA))
    print("=" * LARGURA)

    diagnosticar_ambiente()
    diagnosticar_agente()

    caminhos = [Path(a) for a in sys.argv[1:]]
    if not caminhos:
        titulo("NENHUM ARQUIVO INFORMADO")
        print("  Rode de novo passando um arquivo protegido, por exemplo:")
        print("    python scripts\\diagnostico_drm.py C:\\temp\\planilha.xlsx")
        return 1

    resultados = [diagnosticar_arquivo(c) for c in caminhos]
    veredito(resultados)
    print()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
